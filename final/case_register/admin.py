from django.contrib import admin
# Register your models here.
from django.http import HttpResponseRedirect,HttpResponse
from .models import *
from .resources import *
import openpyxl
from openpyxl.utils import get_column_letter
from import_export.admin import ImportExportModelAdmin
from import_export import resources
from rangefilter.filter import DateRangeFilter, DateTimeRangeFilter
#def export_xlsx(modeladmin, request, queryset):

 #   response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
  #  response['Content-Disposition'] = 'attachment; filename=mymodel.xlsx'
 #   wb = openpyxl.Workbook()
   # ws = wb.get_active_sheet()
   # ws.title = "Case"
#
   # row_num = 0
'''
    columns = [
        (u"cc_date", 15),
        (u"sourcelist", 70),
        (u"case_id", 70),
        (u"customer_name", 70),
        (u"customer_number", 70),
        (u"alt_number", 70),
        (u"companylist", 70),
        (u"vehicle_number", 70),
        (u"vehicle_type", 70),
        (u"payment_type", 70),
        (u"company_cash", 70),
        (u"case_mode", 70),
        (u"amount", 70),
        (u"pay_date", 70),
        (u"vehicle_from", 70),
        (u"location", 70),
        (u"company_kilometer", 70),
        (u"surveyorlist", 70),
        (u"inspection_type", 70),
        (u"surveyor_kilometer", 70),
        (u"surveyor_cash", 70),
        (u"status", 70),
        (u"cd_date", 70),
        (u"time", 70),
        (u"remarks", 70),
    ]

    for col_num in xrange(len(columns)):
        c = ws.cell(row=row_num + 1, column=col_num + 1)
        c.value = columns[col_num][0]
        c.style.font.bold = True
        # set column width
        ws.column_dimensions[get_column_letter(col_num+1)].width = columns[col_num][1]

    for obj in queryset:
        row_num += 1
        row = [
                obj.pk,
                obj.cc_date,
                obj.source_list,
                obj.case_id,
                obj.customer_name,
                obj.customer_number,
                obj.alt_number,
                obj.companylist,
                obj.vehicle_number,
                obj.vehicle_type,
                obj.payment_type,
                obj.company_cash,
                obj.case_mode,
                obj.amount,
                obj.pay_date,
                obj.vehicle_from,
                obj.location,
                obj.company_kilometer,
                obj.surveyorlist,
                obj.inspection_type,
                obj.surveyor_kilometer,
                obj.surveyor_cash,
                obj.status,
                obj.cd_date,
                obj.time,
                obj.remarks,
                obj.source_list,
        ]
        for col_num in xrange(len(row)):
            c = ws.cell(row=row_num + 1, column=col_num + 1)
            c.value = row[col_num]
            c.style.alignment.wrap_text = True

    wb.save(response)
    return response

export_xlsx.short_description = u"Export XLSX"
'''

@admin.register(Surveyorlist)
class listadmin(admin.ModelAdmin):
  fields = ['name']
  list_display = ['name']
  search_fields = ['name']
# @admin.register(Case)
class XLadmin(ImportExportModelAdmin):
    #fields = ['name']
    resource_class = CaseResource
    #fields = ['cc_date','sourcelist','case_id','customer_name','customer_number','alt_number','companylist','vehicle_number','vehicle_type','payment_type','company_cash','case_mode','amount','pay_date','vehicle_from','location','company_kilometer','surveyorlist__name','inspection_type','surveyor_kilometer','surveyor_cash','cd_date','time','remarks','status']
    search_fields = ('vehicle_number','location','case_id')
    list_display = ('cc_date','sourcelist','case_id','vehicle_number','location','surveyorlist','status','remarks')
    list_filter = (('cd_date', DateRangeFilter),'sourcelist','status','surveyorlist__name',)
    #class CaseAdmin(admin.ModelAdmin):
    #   fields = ['cc_date','sourcelist','case_id','customer_name','customer_number','alt_number','companylist','vehicle_number','vehicle_type','payment_type','company_cash','case_mode','amount','pay_date','vehicle_from','location','company_kilometer','surveyorlist','inspection_type','surveyor_kilometer','surveyor_cash','cd_date','time','remarks','status']
    #  search_fields = ('vehicle_number','location')
    # list_display = ('cc_date','sourcelist','case_id','vehicle_number','location','surveyorlist','status','remarks')
        #list_display = ('cc_date','sourcelist','case_id','customer_name','customer_number','alt_number','companylist','vehicle_number','vehicle_type','payment_type','company_cash','case_mode','amount','pay_date','vehicle_from','location','company_kilometer','surveyorlist','inspection_type','surveyor_kilometer','surveyor_cash','cd_date','time','remarks','status')
        #list_filter = ('cc_date','sourcelist','status',)
    # actions = [export_xlsx]




    #list_filter = ('cc_date','sourcelist','status',)
admin.site.register(Case, XLadmin)

#admin.site.register(XLadmin)
#admin.site.register(Surveyorlist)